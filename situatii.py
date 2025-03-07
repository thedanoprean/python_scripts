import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from fuzzywuzzy import fuzz
from collections import defaultdict
import re
import unicodedata

BASE_DIR = r"C:\Users\admitere.DESKTOP-ECSQHO6\Desktop\Situatii Judete"

# Liste de orașe pentru eliminare
ORASE = ["ALBA IULIA", "CÂMPENI", "SEBEȘ", "OCNA MUREȘ", "BLAJ"]

def normalize_text(text):
    """ Normalizează textul pentru comparare (elimină diacritice, ghilimele, caractere speciale). """
    if not text:
        return ""
    
    text = unicodedata.normalize("NFKD", text)  
    text = re.sub(r'[“”„",]', '', text)  # Elimină ghilimelele și virgulele
    text = re.sub(r'\s+', ' ', text).strip()  
    text = text.upper()  
    return text

def remove_city_names(school_name):
    """ Elimină numele orașelor din licee pentru comparare mai bună. """
    for oras in ORASE:
        school_name = school_name.replace(oras, "").strip()
    return school_name

def is_similar_school(name1, name2):
    """ Verifică dacă două licee sunt similare, inclusiv cazuri speciale """

    # Elimină ghilimelele, apostrofurile și normalizează textul
    name1_clean = re.sub(r"[\"'’”“„]", '', name1)
    name2_clean = re.sub(r"[\"'’”“„]", '', name2)

    # Elimină orașele din nume
    name1_clean = remove_city_names(name1_clean)
    name2_clean = remove_city_names(name2_clean)

    # Normalizează numele pentru comparație mai precisă
    name1_clean = re.sub(r'\b(NAT|NATIONAL|COLEGIUL|LICEUL|TEHNOLOGIC|TEORETIC|INDUSTRIAL)\b', '', name1_clean).strip()
    name2_clean = re.sub(r'\b(NAT|NATIONAL|COLEGIUL|LICEUL|TEHNOLOGIC|TEORETIC|INDUSTRIAL)\b', '', name2_clean).strip()

    # Elimină dublările accidentale
    name1_clean = re.sub(r'\s+', ' ', name1_clean).strip()
    name2_clean = re.sub(r'\s+', ' ', name2_clean).strip()

    # Verifică similaritatea normală
    if fuzz.ratio(name1_clean, name2_clean) > 85 or fuzz.partial_ratio(name1_clean, name2_clean) > 90:
        return True

    # Tratare specială pentru liceele de arte
    if ("MUZICA" in name1_clean and "ARTE" in name1_clean) or ("MUZICA" in name2_clean and "ARTE" in name2_clean):
        if "ARTE" in name1_clean and "ARTE" in name2_clean:
            return True  # Se consideră echivalente dacă ambele conțin "ARTE"

    # Tratare specială pentru Colegiul Național Avram Iancu Cîmpeni
    if "AVRAM IANCU" in name1_clean and "AVRAM IANCU" in name2_clean:
        return True

    # Dacă unul este subșir al celuilalt
    if name1_clean in name2_clean or name2_clean in name1_clean:
        return True

    return False

def read_excel_with_bold_marking(file_path, sheet_name):
    """ Citește datele din fișierul Excel și identifică județele marcate cu bold. """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    data = []
    current_judet = None
    
    for row in sheet.iter_rows():
        cell = row[0]
        if cell.font and cell.font.bold:
            current_judet = normalize_text(cell.value)
        else:
            if current_judet and cell.value:
                liceu = normalize_text(cell.value)
                nr_elevi = row[1].value if len(row) > 1 and row[1].value else 0
                data.append((current_judet, liceu, nr_elevi))
    
    return pd.DataFrame(data, columns=["Judet", "Liceu", "Nr. Elevi"])

def group_similar_schools(df):
    """ Grupează liceele similare doar în cadrul aceluiași județ. """
    grouped_data = defaultdict(int)
    
    for judet in df["Judet"].unique():
        df_judet = df[df["Judet"] == judet]

        for _, row in df_judet.iterrows():
            liceu, nr_elevi = row["Liceu"], row["Nr. Elevi"]
            found_match = False

            for existing_liceu in list(grouped_data.keys()):
                existing_judet, existing_liceu_name = existing_liceu

                if existing_judet != judet:
                    continue  

                if is_similar_school(existing_liceu_name, liceu):
                    grouped_data[existing_liceu] += nr_elevi
                    found_match = True
                    break

            if not found_match:
                grouped_data[(judet, liceu)] = nr_elevi

    df_result = pd.DataFrame([(k[0], k[1], v) for k, v in grouped_data.items()], columns=["Judet", "Liceu", "Nr. Elevi"])
    
    total_per_judet = df_result.groupby("Judet")["Nr. Elevi"].sum().to_dict()
    df_result["Total Județ"] = df_result["Judet"].map(total_per_judet)

    return df_result

def save_to_excel(df, folder, filename):
    """ Salvează datele într-un fișier Excel, cu județele grupate și totalul pe județ într-o singură celulă merged. """
    output_path = os.path.join(BASE_DIR, folder, "Rezultat_" + filename)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Rezultat", index=False)
        ws = writer.sheets["Rezultat"]

        current_judet = None
        start_row = 2

        for row in range(2, len(df) + 2):
            if df.iloc[row - 2, 0] != current_judet:
                if current_judet is not None:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                    ws.merge_cells(start_row=start_row, start_column=4, end_row=row - 1, end_column=4)
                current_judet = df.iloc[row - 2, 0]
                start_row = row

    print(f"✅ Fișierul salvat cu succes: {output_path}")

# ---- EXECUȚIA SCRIPTULUI ----
if __name__ == "__main__":
    folder = input("📂 Introdu folderul: ").strip()
    filename = input("📄 Introdu numele fișierului: ").strip()

    file_path = os.path.join(BASE_DIR, folder, filename)

    if os.path.exists(file_path):
        df = read_excel_with_bold_marking(file_path, sheet_name="Foaie2")
        df_grouped = group_similar_schools(df)
        save_to_excel(df_grouped, folder, filename)
    else:
        print("❌ Fișierul nu există!")
